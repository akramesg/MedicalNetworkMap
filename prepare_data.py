import sys
import os
import re
import json
import random
import openpyxl

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

# Source spreadsheet path
EXCEL_PATH = r"c:\Dev\MedicalNetwork\MedSure Network A++ 14-5-2026.xlsx"
OUTPUT_JSON_PATH = r"c:\Dev\MedicalNetwork\providers.json"

# Comprehensive geocoding fallback dictionary for Egypt Governorates and Cities/Districts
FALLBACK_COORDS = {
    # Governorates Centers
    "القاهرة": (30.0444, 31.2357),
    "الجيزة": (30.0131, 31.2089),
    "الاسكندرية": (31.2001, 29.9187),
    "القليوبية": (30.4111, 31.1811), # Benha
    "الغربية": (30.7865, 31.0004), # Tanta
    "الشرقية": (30.5877, 31.5019), # Zagazig
    "الدقهلية": (31.0409, 31.3785), # Mansoura
    "البحيرة": (30.9320, 30.6400), # Damanhour
    "المنوفية": (30.5630, 31.0019), # Shibin El Kom
    "كفر الشيخ": (31.1107, 30.9388),
    "دمياط": (31.4175, 31.8144),
    "بورسعيد": (31.2653, 32.3019),
    "السويس": (29.9668, 32.5498),
    "الاسماعيلية": (30.6044, 32.2722),
    "الفيوم": (29.3084, 30.8428),
    "بني سويف": (29.0744, 31.0978),
    "المنيا": (28.0871, 30.7618),
    "أسيوط": (27.1810, 31.1837),
    "اسيوط": (27.1810, 31.1837),
    "سوهاج": (26.5591, 31.6957),
    "قنا": (26.1551, 32.7160),
    "الاقصر": (25.6872, 32.6396),
    "أسوان": (24.0889, 32.8998),
    "مرسى مطروح": (31.3547, 27.2373),
    "سيناء": (29.5000, 33.8000),
    "جنوب سيناء": (28.5000, 34.0000),
    "شمال سيناء": (30.8000, 33.8000),
    "البحر الاحمر": (25.5000, 34.3000),
    "الوادي الجديد": (25.4390, 30.5486),

    # Specific Major Cities & Districts
    "مصر الجديدة": (30.0910, 31.3235),
    "مدينة نصر": (30.0626, 31.3353),
    "التجمع": (30.0074, 31.4913),
    "التجمع الخامس": (30.0074, 31.4913),
    "المعادى": (29.9602, 31.2569),
    "المهندسين": (30.0617, 31.2013),
    "الدقى": (30.0384, 31.2110),
    "الدقي": (30.0384, 31.2110),
    "وسط البلد": (30.0478, 31.2386),
    "حلوان": (29.8413, 31.3283),
    "شبرا": (30.0818, 31.2483),
    "حدائق القبة": (30.0906, 31.2941),
    "الهرم": (30.0097, 31.1571),
    "فيصل": (30.0167, 31.1333),
    "العبور": (30.2078, 31.4722),
    "الشروق": (30.1472, 31.6319),
    "الشيخ زايد": (30.0175, 31.0028),
    "6 اكتوبر": (29.9722, 30.9419),
    "6 أكتوبر": (29.9722, 30.9419),
    "المقطم": (30.0181, 31.3039),
    "العباسية": (30.0667, 31.2833),
    "سيدي بشر": (31.2644, 30.0089),
    "ميامي": (31.2694, 30.0194),
    "لوران": (31.2494, 29.9822),
    "زيزينيا": (31.2433, 29.9728),
    "سموحة": (31.2089, 29.9634),
    "الغردقة": (27.2579, 33.8116),
    "الجونة": (27.3941, 33.6784),
    "شرم الشيخ": (27.9158, 34.3299),
    "العلمين": (30.8358, 28.9542),
    "طنطا": (30.7865, 31.0004),
    "المحلة الكبرى": (30.9733, 31.1667),
    "المنصورة": (31.0409, 31.3785),
    "الزقازيق": (30.5877, 31.5019),
    "بنها": (30.4594, 31.1856),
    "دمنهور": (30.9320, 30.6400),
    "كفر الدوار": (31.1333, 30.1333),
    "الخارجة": (25.4390, 30.5486),
}

# Compiled Regular Expressions for performance
re_3d = re.compile(r'!3d([0-9.-]+)!4d([0-9.-]+)')
re_1d2d = re.compile(r'!1d([0-9.-]+)!2d([0-9.-]+)')
re_dir_double = re.compile(r'/dir//([0-9.-]+),([0-9.-]+)')
re_dir_single = re.compile(r'/dir/[^/]+/([0-9.-]+),([0-9.-]+)')
re_query = re.compile(r'query=([0-9.-]+),([0-9.-]+)')
re_at = re.compile(r'@([0-9.-]+),([0-9.-]+)')

def extract_coords(url):
    if not url:
        return None
        
    # 1. Try exact place pin (!3d/!4d)
    m = re_3d.search(url)
    if m:
        try:
            lat, lng = float(m.group(1)), float(m.group(2))
            if 20 <= lat <= 33 and 23 <= lng <= 38:
                return lat, lng, "exact_place_3d"
        except ValueError:
            pass

    # 2. Try exact directions destination parameter (!1d...!2d)
    # The last occurrence is always the destination coords
    matches = re_1d2d.findall(url)
    if matches:
        try:
            lng_str, lat_str = matches[-1]
            lng, lat = float(lng_str), float(lat_str)
            if 20 <= lat <= 33 and 23 <= lng <= 38:
                return lat, lng, "dir_dest_1d2d"
        except ValueError:
            pass

    # 3. Try directions destination coordinate: /dir//lat,lng
    m = re_dir_double.search(url)
    if m:
        try:
            lat, lng = float(m.group(1)), float(m.group(2))
            if 20 <= lat <= 33 and 23 <= lng <= 38:
                return lat, lng, "dir_dest_double"
        except ValueError:
            pass
            
    # 4. Try directions single slash coordinate: /dir/Address/lat,lng
    m = re_dir_single.search(url)
    if m:
        try:
            lat, lng = float(m.group(1)), float(m.group(2))
            if 20 <= lat <= 33 and 23 <= lng <= 38:
                return lat, lng, "dir_dest_single"
        except ValueError:
            pass

    # 5. Try query coordinate parameter: query=lat,lng
    m = re_query.search(url)
    if m:
        try:
            lat, lng = float(m.group(1)), float(m.group(2))
            if 20 <= lat <= 33 and 23 <= lng <= 38:
                return lat, lng, "query_coord"
        except ValueError:
            pass

    # 6. Try viewport coordinate: @lat,lng (strictly as a viewport fallback)
    m = re_at.search(url)
    if m:
        try:
            lat, lng = float(m.group(1)), float(m.group(2))
            if 20 <= lat <= 33 and 23 <= lng <= 38:
                return lat, lng, "viewport_fallback_at"
        except ValueError:
            pass
            
    return None

def extract_coords_from_address(addr):
    if not addr:
        return None
    # 1. Search for any Google Maps URL in the address
    urls = re.findall(r'(https?://[^\s]+)', addr)
    for u in urls:
        res = extract_coords(u)
        if res:
            return res[0], res[1], "addr_url_" + res[2]
            
    # 2. Search for explicit decimal coordinate pairs in address, e.g. 30.012345, 31.123456
    coords_match = re.search(r'([23]\d\.\d{4,8})\s*,\s*([23]\d\.\d{4,8})', addr)
    if coords_match:
        try:
            lat, lng = float(coords_match.group(1)), float(coords_match.group(2))
            if 20 <= lat <= 33 and 23 <= lng <= 38:
                return lat, lng, "addr_explicit_coords"
        except ValueError:
            pass
            
    return None

def clean_str(val):
    if val is None:
        return ""
    s = str(val).strip()
    s = s.replace("\xa0", " ").strip()
    return s

def main():
    print("Loading workbook...")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    
    sheetname = 'جميع المحافظات'
    if sheetname not in wb.sheetnames:
        print(f"Error: Sheet '{sheetname}' not found!")
        return
        
    sheet = wb[sheetname]
    print(f"Reading sheet '{sheetname}'...")
    
    rows_iter = sheet.iter_rows(values_only=True)
    headers = next(rows_iter)
    
    try:
        gov_idx = headers.index('المحافظة')
        city_idx = headers.index('المدينة/الحي')
        type_idx = headers.index('نوع مقدم الخدمة')
        name_idx = headers.index('إسم مقدم الخدمة')
        class_idx = headers.index('تصنيف الشبكة')
        status_idx = headers.index('الحالة')
        spec_idx = headers.index('التخصصات')
        serv_idx = headers.index('الخدمات المؤداه')
        addr_idx = headers.index('العنوان')
        loc_idx = headers.index('الموقع')
        hours_idx = headers.index('مواعيد العمل')
        sched_idx = headers.index('جدول مواعيد العيادات')
    except ValueError as e:
        print("Error: Could not find required column in sheet headers:", e)
        return

    providers = []
    coord_stats = {
        "exact_place_3d": 0,
        "dir_dest_1d2d": 0,
        "dir_dest_double": 0,
        "dir_dest_single": 0,
        "query_coord": 0,
        "viewport_fallback_at": 0,
        "addr_url_exact_place_3d": 0,
        "addr_url_dir_dest_1d2d": 0,
        "addr_url_dir_dest_double": 0,
        "addr_url_dir_dest_single": 0,
        "addr_url_query_coord": 0,
        "addr_url_viewport_fallback_at": 0,
        "addr_explicit_coords": 0,
        "fallback_addr_neighborhood": 0,
        "fallback_city": 0,
        "fallback_gov": 0,
        "fallback_egypt": 0
    }
    
    row_count = 0
    for r in rows_iter:
        if not any(r):
            continue
        
        row_count += 1
        gov = clean_str(r[gov_idx])
        city = clean_str(r[city_idx])
        ptype = clean_str(r[type_idx])
        name = clean_str(r[name_idx])
        pclass = clean_str(r[class_idx])
        spec = clean_str(r[spec_idx])
        serv = clean_str(r[serv_idx])
        addr = clean_str(r[addr_idx])
        url = clean_str(r[loc_idx])
        hours = clean_str(r[hours_idx])
        sched = clean_str(r[sched_idx])

        if not name:
            continue
            
        coords_res = extract_coords(url)
        lat, lng = None, None
        method = ""
        
        if coords_res:
            lat, lng, method = coords_res
            coord_stats[method] += 1
        else:
            # 1. Try to extract exact coordinates from the Address field!
            addr_coords = extract_coords_from_address(addr)
            if addr_coords:
                lat, lng, method = addr_coords
                if method not in coord_stats:
                    coord_stats[method] = 0
                coord_stats[method] += 1
            else:
                # Geocoding fallbacks using neighborhoods, city or governorates
                addr_normalized = addr.replace("أ", "ا").replace("إ", "ا").replace("ة", "ه").strip()
                found = False
                
                # A. Try to match specific neighborhoods or major cities in the Address field first!
                specific_neighborhoods = [
                    "مصر الجديدة", "مدينة نصر", "التجمع الخامس", "التجمع", "المعادى", "المهندسين", 
                    "الدقى", "الدقي", "وسط البلد", "حلوان", "شبرا", "حدائق القبة", "الهرم", "فيصل", 
                    "العبور", "الشروق", "الشيخ زايد", "6 اكتوبر", "6 أكتوبر", "المقطم", "العباسية", 
                    "سيدي بشر", "ميامي", "لوران", "زيزينيا", "سموحة", "الغردقة", "الجونة", "شرم الشيخ", "العلمين"
                ]
                for k in sorted(FALLBACK_COORDS.keys(), key=len, reverse=True):
                    k_normalized = k.replace("أ", "ا").replace("إ", "ا").replace("ة", "ه").strip()
                    if k in specific_neighborhoods and k_normalized in addr_normalized:
                        lat, lng = FALLBACK_COORDS[k]
                        coord_stats["fallback_addr_neighborhood"] += 1
                        found = True
                        break
                
                # B. Match specific city
                if not found:
                    city_normalized = city.replace("أ", "ا").replace("إ", "ا").replace("ة", "ه").strip()
                    for k, (clat, clng) in FALLBACK_COORDS.items():
                        k_normalized = k.replace("أ", "ا").replace("إ", "ا").replace("ة", "ه").strip()
                        if city_normalized and (city_normalized in k_normalized or k_normalized in city_normalized):
                            lat, lng = clat, clng
                            coord_stats["fallback_city"] += 1
                            found = True
                            break
                
                # C. Match Governorate
                if not found:
                    gov_normalized = gov.replace("أ", "ا").replace("إ", "ا").replace("ة", "ه").strip()
                    for k, (glat, glng) in FALLBACK_COORDS.items():
                        k_normalized = k.replace("أ", "ا").replace("إ", "ا").replace("ة", "ه").strip()
                        if gov_normalized and (gov_normalized in k_normalized or k_normalized in gov_normalized):
                            lat, lng = glat, glng
                            coord_stats["fallback_gov"] += 1
                            found = True
                            break
                
                # D. Fallback to Cairo
                if not found:
                    lat, lng = FALLBACK_COORDS["القاهرة"]
                    coord_stats["fallback_egypt"] += 1
                
                # Jitter slightly for visual scatter
                lat += random.uniform(-0.008, 0.008)
                lng += random.uniform(-0.008, 0.008)

        if not pclass:
            pclass = "A"

        provider_obj = {
            "id": row_count,
            "g": gov,
            "c": city,
            "t": ptype,
            "n": name,
            "nc": pclass,
            "sp": spec,
            "se": serv,
            "a": addr,
            "lat": round(lat, 6),
            "lng": round(lng, 6),
            "u": url,
            "h": hours,
            "sc": sched
        }
        
        providers.append(provider_obj)

    print(f"\nProcessed {len(providers)} providers successfully!")
    print("Coordinate Extraction Statistics:")
    for k, v in coord_stats.items():
        pct = (v / len(providers)) * 100
        print(f" - {k}: {v} ({pct:.1f}%)")

    total_exact = sum(v for k, v in coord_stats.items() if not k.startswith("fallback"))
    total_fallback = sum(v for k, v in coord_stats.items() if k.startswith("fallback"))
    print(f"\nTotal exact coordinates resolved: {total_exact} ({total_exact/len(providers)*100:.1f}%)")
    print(f"Total fallback regional geocoded: {total_fallback} ({total_fallback/len(providers)*100:.1f}%)")

    # Export to compact JSON
    print(f"Saving to '{OUTPUT_JSON_PATH}'...")
    with open(OUTPUT_JSON_PATH, 'w', encoding='utf-8') as f:
        json.dump(providers, f, ensure_ascii=False, separators=(',', ':'))
        
    print(f"Data pre-processing completed! File size: {os.path.getsize(OUTPUT_JSON_PATH)/1024/1024:.2f} MB")

if __name__ == "__main__":
    main()
